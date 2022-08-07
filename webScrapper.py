import requests 
from bs4 import BeautifulSoup
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = "Top Rated Movies"
print(excel.sheetnames)
sheet.append(['Rank', 'Name', 'Year of Release', 'Ratings'])


#source is response object
try:
    source = requests.get("https://www.imdb.com/chart/top/")
    source.raise_for_status()#throw error 404 if url not found
    
    #returns BS obj and stores in var soup
    soup = BeautifulSoup(source.text,'html.parser')#pass text of the response object
    
    #find used to find 1st match; _ imp for finding class
    movie = soup.find('tbody', class_='lister-list').find_all('tr')
    
    #print(len(movie))
    
    #iterate through each <tr> tag
    for mov in movie:
        
        #entering tag and extract name of movie
        name = mov.find('td', class_='titleColumn').a.text
        #to get rank of movie and split "."
        rank = mov.find('td', class_='titleColumn').get_text(strip=True).split(".")[0]
        #access year
        year = mov.find('td', class_='titleColumn').span.text.strip('()')
        #access rating
        rating = mov.find('td', class_='ratingColumn imdbRating').strong.text
        
        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])
        #break
    
except Exception as e:
    print(e)

excel.save('Movie Ratings.xlsx')